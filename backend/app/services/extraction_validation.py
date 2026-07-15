"""
Extraction completeness validation (pre-structuring).

Verifies that the lossless extraction faithfully represents the workbook by
INDEPENDENTLY recounting the workbook straight from openpyxl — deliberately
NOT reusing lossless_extractor's code — and comparing the two. If the two
disagree, the extractor dropped or miscounted something, and we want to know
before spending any structuring effort.

Also audits legend usage against the transformation-engine intermediate:
every detected legend colour/code should actually match cells in the sheet,
and every colour that carries data should ideally have a legend meaning.

Entirely deterministic and generic.
"""
import openpyxl


def _independent_sheet_counts(worksheet):
    """A from-scratch recount using openpyxl directly (no lossless code)."""
    min_row = worksheet.min_row or 1
    max_row = worksheet.max_row or 0
    min_col = worksheet.min_column or 1
    max_col = worksheet.max_column or 0

    total = value = blank = colored = 0
    comment_cells = hyperlink_cells = 0
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = worksheet.cell(row=r, column=c)
            total += 1
            v = cell.value
            if v is None or v == "":
                blank += 1
            else:
                value += 1
            fill = getattr(cell, "fill", None)
            if fill is not None and getattr(fill, "patternType", None) is not None:
                colored += 1
            if getattr(cell, "comment", None) is not None:
                comment_cells += 1
            hl = getattr(cell, "hyperlink", None)
            if hl is not None and getattr(hl, "target", None) is not None:
                hyperlink_cells += 1

    hidden_rows = sum(1 for _, d in worksheet.row_dimensions.items() if d.hidden)
    hidden_cols = sum(1 for _, d in worksheet.column_dimensions.items() if d.hidden)
    return {
        "total_cells": total,
        "value_cells": value,
        "blank_cells": blank,
        # NOTE: this counts cells with ANY pattern fill; the lossless
        # "colored_cells" counts only cells whose fill resolves to a hex, so
        # these two are compared with that distinction in mind below.
        "filled_cells": colored,
        "comment_cells": comment_cells,
        "hyperlink_cells": hyperlink_cells,
        "merged_ranges": len(list(worksheet.merged_cells.ranges)),
        "hidden_rows": hidden_rows,
        "hidden_columns": hidden_cols,
    }


def validate_extraction_completeness(file_path, lossless, intermediate=None):
    """
    Returns a report:
      {
        "ok": bool,
        "sheet_count": {"workbook": n, "lossless": n, "match": bool},
        "sheet_names_match": bool,
        "per_sheet": {name: {"match": bool, "checks": {...}, "issues": [...]}},
        "legend_usage": {name: [{color/code, matched_cells, ...}]},
        "issues": [...top-level...]
      }
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    report = {"ok": True, "per_sheet": {}, "legend_usage": {}, "issues": []}

    wb_names = list(wb.sheetnames)
    loss_names = list(lossless.get("sheets", {}).keys())
    report["sheet_count"] = {
        "workbook": len(wb_names),
        "lossless": len(loss_names),
        "match": len(wb_names) == len(loss_names),
    }
    report["sheet_names_match"] = set(wb_names) == set(loss_names)
    if not report["sheet_names_match"]:
        report["ok"] = False
        report["issues"].append(
            f"sheet names differ: workbook={wb_names} lossless={loss_names}"
        )

    for name in wb_names:
        loss_sheet = lossless.get("sheets", {}).get(name)
        if loss_sheet is None:
            report["ok"] = False
            report["issues"].append(f"sheet '{name}' present in workbook but missing from lossless extraction")
            continue

        indep = _independent_sheet_counts(wb[name])
        lv = loss_sheet.get("validation", {})
        checks = {}
        issues = []

        # Cells: must match exactly (both count the full used-range rectangle).
        checks["total_cells"] = {"workbook": indep["total_cells"], "lossless": lv.get("total_cells"), "match": indep["total_cells"] == lv.get("total_cells")}
        checks["value_cells"] = {"workbook": indep["value_cells"], "lossless": lv.get("value_cells"), "match": indep["value_cells"] == lv.get("value_cells")}
        checks["blank_cells"] = {"workbook": indep["blank_cells"], "lossless": lv.get("blank_cells"), "match": indep["blank_cells"] == lv.get("blank_cells")}
        checks["merged_ranges"] = {"workbook": indep["merged_ranges"], "lossless": lv.get("merged_ranges"), "match": indep["merged_ranges"] == lv.get("merged_ranges")}
        checks["comment_cells"] = {"workbook": indep["comment_cells"], "lossless": lv.get("comment_cells"), "match": indep["comment_cells"] == lv.get("comment_cells")}
        checks["hyperlink_cells"] = {"workbook": indep["hyperlink_cells"], "lossless": lv.get("hyperlink_cells"), "match": indep["hyperlink_cells"] == lv.get("hyperlink_cells")}
        checks["hidden_rows"] = {"workbook": indep["hidden_rows"], "lossless": lv.get("hidden_rows"), "match": indep["hidden_rows"] == lv.get("hidden_rows")}
        checks["hidden_columns"] = {"workbook": indep["hidden_columns"], "lossless": lv.get("hidden_columns"), "match": indep["hidden_columns"] == lv.get("hidden_columns")}
        # Colored: lossless resolves to a hex, so lossless.colored <= workbook.filled.
        colored_ok = (lv.get("colored_cells") or 0) <= indep["filled_cells"]
        checks["colored_cells"] = {
            "workbook_any_fill": indep["filled_cells"],
            "lossless_resolved": lv.get("colored_cells"),
            "match": colored_ok,
        }

        for key, chk in checks.items():
            if not chk["match"]:
                issues.append(f"{key}: {chk}")

        sheet_ok = not issues
        if not sheet_ok:
            report["ok"] = False
        report["per_sheet"][name] = {"match": sheet_ok, "checks": checks, "issues": issues}

    # Legend usage audit (needs the transformation-engine intermediate).
    if intermediate is not None:
        for name, inter in intermediate.get("sheets", {}).items():
            report["legend_usage"][name] = _legend_usage(inter)

    wb.close()
    return report


def _legend_usage(intermediate_sheet):
    """
    For each legend entry, count how many cells in the sheet actually match it
    (item 16). Also flag data colours that have no legend meaning.
    """
    color_counts = {e["hex"]: e["count"] for e in intermediate_sheet.get("color_map", [])}
    used_colors = set()
    used_codes = set()
    usage = []

    for cand in intermediate_sheet.get("legend_candidates", []):
        hexv = cand["color_hex"]
        matched = color_counts.get(hexv, 0)
        used_colors.add(hexv)
        if cand["code"]:
            used_codes.add(cand["code"])
        usage.append(
            {
                "kind": cand["kind"],
                "color_hex": hexv,
                "code": cand["code"],
                "label": cand["label"],
                "matched_cells": matched,
                "ignored": matched == 0,
            }
        )

    # Colours that carry data but were never explained by a legend entry.
    unexplained = [
        {"color_hex": hexv, "count": cnt}
        for hexv, cnt in color_counts.items()
        if hexv not in used_colors and hexv not in ("000000", "FFFFFF")
    ]
    return {"entries": usage, "unexplained_colors": unexplained}
