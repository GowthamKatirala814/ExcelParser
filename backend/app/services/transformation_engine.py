"""
Deterministic Transformation Engine.

Converts one sheet's *lossless* raw JSON (services/lossless_extractor.py) into
an intermediate, business-aware JSON — WITHOUT any LLM. Everything here is
deterministic Python: it detects table boundaries, header rows, horizontal
date headers, the distinct-colour inventory, and legend candidates
(colour/code -> label), then emits row objects in which every value is
traceable back to its exact Excel cell.

The point of this layer is that Gemini never has to rediscover structure:
by the time a sheet reaches the LLM, tables / dates / colours / legends are
already found. Gemini's job shrinks to naming and organising.

Nothing here is workbook-, supplier-, or layout-specific. It consumes only
the generic lossless representation, so it runs identically on any workbook.
"""
import datetime

from openpyxl.utils import get_column_letter, range_boundaries


# --------------------------------------------------------------------------
# Grid helpers over the lossless cell list
# --------------------------------------------------------------------------


def _build_grid(sheet):
    """(row, col) -> lossless cell dict, for every emitted cell."""
    return {(c["row"], c["column"]): c for c in sheet.get("cells", [])}


def _has_value(cell):
    if cell is None:
        return False
    v = cell.get("value")
    return v is not None and v != ""


def _fill_hex(cell):
    if cell is None:
        return None
    fill = cell.get("fill")
    if not fill:
        return None
    return fill.get("resolved_hex")


def _is_bold(cell):
    if cell is None:
        return False
    font = cell.get("font")
    return bool(font and font.get("bold"))


def _parse_date(cell):
    """Return a datetime.date if the cell holds a date, else None. Lossless
    serialises datetimes to ISO strings and tags them data_type 'd'."""
    if cell is None:
        return None
    value = cell.get("value")
    if not isinstance(value, str):
        return None
    if cell.get("data_type") != "d" and "T" not in value and "-" not in value:
        return None
    try:
        return datetime.date.fromisoformat(value[:10])
    except ValueError:
        return None


def _is_int(cell):
    if cell is None:
        return False
    v = cell.get("value")
    return isinstance(v, int) and not isinstance(v, bool)


# --------------------------------------------------------------------------
# Table / header / date-column detection (deterministic, layout-only)
# --------------------------------------------------------------------------


def _detect_blocks(grid, min_row, max_row, min_col, max_col):
    """Split rows into blocks separated by one or more fully-blank rows."""
    blocks = []
    start = None
    end = None
    for r in range(min_row, max_row + 1):
        row_has = any(_has_value(grid.get((r, c))) for c in range(min_col, max_col + 1))
        if row_has:
            if start is None:
                start = r
            end = r
        elif start is not None:
            blocks.append((start, end))
            start = None
    if start is not None:
        blocks.append((start, end))
    return blocks


def _row_pattern(grid, r, min_col, max_col):
    return tuple(_has_value(grid.get((r, c))) for c in range(min_col, max_col + 1))


def _horizontal_merges_by_row(merged_ranges):
    """row -> True if a merge on that row spans >= 2 columns."""
    out = {}
    for ref in merged_ranges or []:
        try:
            min_c, min_r, max_c, max_r = range_boundaries(ref)
        except (ValueError, TypeError):
            continue
        if min_r == max_r and (max_c - min_c) >= 1:
            out[min_r] = True
    return out


def _detect_header(grid, start, end, min_col, max_col, hmerge_rows):
    is_bold = any(
        _is_bold(grid.get((start, c))) and _has_value(grid.get((start, c)))
        for c in range(min_col, max_col + 1)
    )
    is_merged = start in hmerge_rows
    repeats = False
    if start + 2 <= end:
        p1 = _row_pattern(grid, start + 1, min_col, max_col)
        p2 = _row_pattern(grid, start + 2, min_col, max_col)
        if p1 == p2 and any(p1):
            repeats = True
    if is_bold or is_merged or repeats:
        return start, {"bold": is_bold, "merged_across_columns": is_merged, "repeated_data_row_pattern": repeats}
    return None, None


def _detect_date_row(grid, start, end, min_col, max_col):
    """
    Two positional patterns (same as the compact engine):
      1. a row that is entirely dates across >= 2 columns.
      2. exactly one date (a month anchor) + the integer sequence 1,2,3...
    Returns (date_row, anchor_date) — anchor_date only set for pattern 2.
    """
    for r in range(start, end):
        present = [grid.get((r, c)) for c in range(min_col, max_col + 1) if _has_value(grid.get((r, c)))]
        if len(present) >= 2 and all(_parse_date(c) is not None for c in present):
            return r, None
    for r in range(start, end):
        present = [grid.get((r, c)) for c in range(min_col, max_col + 1) if _has_value(grid.get((r, c)))]
        if len(present) < 3:
            continue
        dates = [c for c in present if _parse_date(c) is not None]
        ints = [c["value"] for c in present if _is_int(c)]
        if len(dates) == 1 and len(ints) == len(present) - 1 and ints == list(range(1, len(ints) + 1)):
            return r, _parse_date(dates[0])
    return None, None


def _build_date_columns(grid, date_row, anchor_date, min_col, max_col):
    index = {}
    if anchor_date is not None:
        for c in range(min_col, max_col + 1):
            cell = grid.get((date_row, c))
            if not _is_int(cell):
                continue
            try:
                d = anchor_date.replace(day=cell["value"])
            except ValueError:
                d = None
            index[get_column_letter(c)] = {
                "cell": cell["cell_ref"],
                "date": d.isoformat() if d else None,
                "day_number": cell["value"],
            }
        return index
    day = 0
    for c in range(min_col, max_col + 1):
        cell = grid.get((date_row, c))
        d = _parse_date(cell)
        if d is None:
            continue
        day += 1
        index[get_column_letter(c)] = {"cell": cell["cell_ref"], "date": d.isoformat(), "day_number": day}
    return index


# --------------------------------------------------------------------------
# Colour inventory + legend detection
# --------------------------------------------------------------------------


def _build_color_map(sheet):
    inv = {}
    for cell in sheet.get("cells", []):
        hexv = _fill_hex(cell)
        if not hexv:
            continue
        entry = inv.setdefault(hexv, {"hex": hexv, "count": 0, "example_cells": []})
        entry["count"] += 1
        if len(entry["example_cells"]) < 8:
            entry["example_cells"].append(cell["cell_ref"])
    return sorted(inv.values(), key=lambda e: e["count"], reverse=True)


def _looks_like_label(cell):
    """A descriptive label: text of length >= 3 containing a letter, not a date."""
    if not _has_value(cell):
        return False
    v = cell.get("value")
    if not isinstance(v, str):
        return False
    if _parse_date(cell) is not None:
        return False
    stripped = v.strip()
    return len(stripped) >= 3 and any(ch.isalpha() for ch in stripped)


def _detect_legend_candidates(grid, min_row, max_row, min_col, max_col):
    """
    A legend entry is a filled 'swatch' cell (blank or a short code) whose
    immediate right neighbour holds descriptive label text. Emits two kinds:
      - fill_meaning: blank swatch + colour  -> label   (the colour's base meaning)
      - code_meaning: short code + colour     -> label   (a code that overrides)
    Conservative by design: the data grid never has coloured cells sitting
    directly left of free-form descriptive text, so false positives are rare.
    Results are candidates; downstream semantic cleanup can confirm them.
    """
    candidates = []
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            swatch = grid.get((r, c))
            hexv = _fill_hex(swatch)
            if not hexv:
                continue
            label_cell = grid.get((r, c + 1))
            if not _looks_like_label(label_cell):
                continue
            code = swatch.get("value")
            code_str = str(code).strip() if code not in (None, "") else None
            if code_str is not None and len(code_str) > 4:
                continue  # too long to be a status code; likely data, skip
            label_text = label_cell["value"].strip()
            if code_str is not None and code_str.lower() == label_text.lower():
                continue
            candidates.append(
                {
                    "kind": "code_meaning" if code_str else "fill_meaning",
                    "color_hex": hexv,
                    "code": code_str,
                    "label": label_text,
                    "swatch_cell": swatch["cell_ref"],
                    "label_cell": label_cell["cell_ref"],
                }
            )
    return candidates


def _build_legend_maps(candidates):
    """Convenience lookups derived from candidates: colour->label (base) and
    code->label (override)."""
    color_to_label = {}
    code_to_label = {}
    for cand in candidates:
        if cand["kind"] == "fill_meaning" and cand["color_hex"] not in color_to_label:
            color_to_label[cand["color_hex"]] = cand["label"]
        if cand["code"] and cand["code"] not in code_to_label:
            code_to_label[cand["code"]] = {"label": cand["label"], "color_hex": cand["color_hex"]}
    return color_to_label, code_to_label


# --------------------------------------------------------------------------
# Row objects with cell-ref traceability
# --------------------------------------------------------------------------


def _build_rows(grid, start, end, min_col, max_col, date_columns):
    rows = []
    date_cols = set(date_columns.keys())
    for r in range(start, end + 1):
        cells = []
        for c in range(min_col, max_col + 1):
            cell = grid.get((r, c))
            if cell is None:
                continue
            letter = get_column_letter(c)
            hexv = _fill_hex(cell)
            if not _has_value(cell) and hexv is None:
                continue  # nothing to carry for this position
            entry = {
                "cell": cell["cell_ref"],
                "column": letter,
                "value": cell.get("value"),
                "fill_hex": hexv,
            }
            if letter in date_cols:
                entry["date"] = date_columns[letter]["date"]
            cells.append(entry)
        if cells:
            rows.append({"row": r, "cells": cells})
    return rows


# --------------------------------------------------------------------------
# Public entry point
# --------------------------------------------------------------------------


def transform_sheet(lossless_sheet):
    """
    lossless_sheet: one sheet from extract_lossless()['sheets'][name].
    Returns the deterministic intermediate JSON for that sheet.
    """
    grid = _build_grid(lossless_sheet)
    dims = lossless_sheet.get("dimensions", {})
    min_row = dims.get("min_row") or 1
    max_row = dims.get("max_row") or 0
    min_col = dims.get("min_col") or 1
    max_col = dims.get("max_col") or 0

    hmerge_rows = _horizontal_merges_by_row(lossless_sheet.get("merged_ranges"))
    color_map = _build_color_map(lossless_sheet)
    legend_candidates = _detect_legend_candidates(grid, min_row, max_row, min_col, max_col)
    color_to_label, code_to_label = _build_legend_maps(legend_candidates)

    tables = []
    for i, (start, end) in enumerate(_detect_blocks(grid, min_row, max_row, min_col, max_col)):
        header_row, signals = _detect_header(grid, start, end, min_col, max_col, hmerge_rows)
        date_row, anchor_date = _detect_date_row(grid, start, end, min_col, max_col)
        date_columns = (
            _build_date_columns(grid, date_row, anchor_date, min_col, max_col)
            if date_row is not None
            else {}
        )
        data_start = start + 1 if header_row is not None else start

        header = None
        if header_row is not None:
            header = {
                "row": header_row,
                "signals": signals,
                "cells": [
                    {"cell": grid[(header_row, c)]["cell_ref"], "column": get_column_letter(c), "value": grid[(header_row, c)].get("value")}
                    for c in range(min_col, max_col + 1)
                    if grid.get((header_row, c)) is not None and _has_value(grid.get((header_row, c)))
                ],
            }

        ambiguous = header_row is None and end > start
        tables.append(
            {
                "table_index": i,
                "row_range": [start, end],
                "header": header,
                "date_columns": date_columns or None,
                "date_column_count": len(date_columns),
                "rows": _build_rows(grid, data_start, end, min_col, max_col, date_columns),
                "ambiguous": ambiguous,
                "ambiguous_reason": (
                    f"block rows {start}-{end} has no bold / merged-header / repeated-shape signal"
                    if ambiguous
                    else None
                ),
            }
        )

    return {
        "sheet": lossless_sheet.get("name"),
        "state": lossless_sheet.get("state"),
        "dimensions": dims,
        "merged_ranges": lossless_sheet.get("merged_ranges", []),
        "color_map": color_map,
        "legend_candidates": legend_candidates,
        "legend": {"color_to_label": color_to_label, "code_to_label": code_to_label},
        "tables": tables,
        "summary": {
            "table_count": len(tables),
            "distinct_colors": len(color_map),
            "legend_entries": len(legend_candidates),
            "tables_with_date_headers": sum(1 for t in tables if t["date_columns"]),
        },
    }


def transform_workbook(lossless):
    """Runs transform_sheet on every sheet in a lossless workbook extraction."""
    return {
        "workbook": lossless.get("workbook", {}),
        "sheets": {name: transform_sheet(sheet) for name, sheet in lossless.get("sheets", {}).items()},
    }
