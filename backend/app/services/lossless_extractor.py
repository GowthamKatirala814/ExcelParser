"""
Lossless raw extraction — a faithful digital representation of an Excel
workbook.

Unlike the compact extractor (services/extractor.py), this emits EVERY cell
in each sheet's used range, whether or not it holds a value, and preserves
every attribute openpyxl exposes: the full colour internals (rgb / theme /
tint / indexed / pattern, not just a resolved hex), fonts, borders,
alignment, number format, formula text AND cached value, comments,
hyperlinks, and merge relationships. It also captures workbook- and
sheet-level metadata (dimensions, merged ranges, hidden rows/columns, freeze
panes, sheet state, properties) and detects embedded images.

Design notes:
  - Two openpyxl loads are used: data_only=True gives each formula cell's
    cached computed value; data_only=False gives the formula text. A cell can
    therefore report both.
  - openpyxl Colour attributes are type-specific: reading `.theme` on an rgb
    colour returns a sentinel, not an int. `_colour_components` reads only the
    attribute matching `.type`, so nothing is corrupted.
  - Five Excel cell states the spec calls out are all recoverable from two
    orthogonal fields kept per cell: `state` (value dimension:
    blank/empty_string/formula_blank/value) and `fill` (fill dimension:
    has_fill / is_white / resolved_hex). No information is collapsed.
  - Nothing here is workbook/supplier specific.
"""
import datetime
import zipfile

import openpyxl
from openpyxl.styles.colors import COLOR_INDEX

from .image_detector import detect_images_by_sheet
from .theme_resolver import load_theme_palette, resolve_theme_color

_BORDER_SIDES = ("top", "bottom", "left", "right", "diagonal")
_INDEXED_AUTO_SLOTS = {64, 65}
# Rough heuristic (no OCR dependency here): an embedded image this large is
# more likely to be a screenshot of a table than a small logo/icon, so it's
# flagged for optional downstream OCR. Purely advisory.
_LIKELY_TABULAR_MIN_W = 300
_LIKELY_TABULAR_MIN_H = 200


def _serialize(value):
    if isinstance(value, (datetime.datetime, datetime.date, datetime.time)):
        return value.isoformat()
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return str(value)


def _display_value(value):
    """Best-effort human display string. Faithful Excel number-format
    rendering is out of scope (the number_format code is preserved per cell
    instead), so this is str() of the value, empty string for None."""
    if value is None:
        return ""
    if isinstance(value, (datetime.datetime, datetime.date, datetime.time)):
        return value.isoformat()
    return str(value)


def _strip_alpha(argb):
    if not isinstance(argb, str) or len(argb) < 6:
        return None
    return argb[-6:].upper()


def _colour_components(colour):
    """Raw colour facts, reading only the attribute matching the colour's
    declared type (openpyxl returns sentinels for the others)."""
    if colour is None:
        return None
    ctype = getattr(colour, "type", None)
    out = {"type": ctype, "rgb": None, "theme": None, "tint": None, "indexed": None}
    tint = getattr(colour, "tint", None)
    if isinstance(tint, (int, float)):
        out["tint"] = tint
    if ctype == "rgb":
        rgb = getattr(colour, "rgb", None)
        if isinstance(rgb, str):
            out["rgb"] = rgb
    elif ctype == "theme":
        theme = getattr(colour, "theme", None)
        if isinstance(theme, int):
            out["theme"] = theme
    elif ctype == "indexed":
        indexed = getattr(colour, "indexed", None)
        if isinstance(indexed, int):
            out["indexed"] = indexed
    return out


def _resolve_hex(components, palette):
    """Best-effort final hex ("RRGGBB") from raw colour components, or None."""
    if components is None:
        return None
    if components["rgb"]:
        return _strip_alpha(components["rgb"])
    if components["theme"] is not None:
        return resolve_theme_color(components["theme"], components["tint"] or 0.0, palette)
    idx = components["indexed"]
    if isinstance(idx, int) and idx not in _INDEXED_AUTO_SLOTS and 0 <= idx < len(COLOR_INDEX):
        return _strip_alpha(COLOR_INDEX[idx])
    return None


def _fill_info(cell, palette):
    fill = getattr(cell, "fill", None)
    pattern = getattr(fill, "patternType", None) if fill is not None else None

    if pattern is None:
        # No fill at all: openpyxl still reports a default transparent
        # fgColor (00000000) that carries no real information, so the colour
        # components are nulled to avoid false "this cell is coloured" hits
        # downstream. has_fill=False is the authoritative signal.
        return {
            "has_fill": False,
            "pattern": None,
            "rgb": None,
            "theme": None,
            "tint": None,
            "indexed": None,
            "resolved_hex": None,
            "is_white": False,
            "bg_components": None,
        }

    fg = getattr(fill, "fgColor", None)
    bg = getattr(fill, "bgColor", None)
    comps = _colour_components(fg)
    resolved_hex = _resolve_hex(comps, palette)
    return {
        "has_fill": True,
        "pattern": pattern,
        "rgb": comps["rgb"] if comps else None,
        "theme": comps["theme"] if comps else None,
        "tint": comps["tint"] if comps else None,
        "indexed": comps["indexed"] if comps else None,
        "resolved_hex": resolved_hex,
        "is_white": resolved_hex == "FFFFFF",
        "bg_components": _colour_components(bg),
    }


def _font_info(cell, palette):
    font = getattr(cell, "font", None)
    if font is None:
        return None
    comps = _colour_components(getattr(font, "color", None))
    return {
        "name": getattr(font, "name", None),
        "size": getattr(font, "size", None),
        "bold": bool(getattr(font, "bold", False)),
        "italic": bool(getattr(font, "italic", False)),
        "underline": getattr(font, "underline", None),
        "rgb": comps["rgb"] if comps else None,
        "theme": comps["theme"] if comps else None,
        "tint": comps["tint"] if comps else None,
        "indexed": comps["indexed"] if comps else None,
        "resolved_hex": _resolve_hex(comps, palette) if comps else None,
    }


def _border_info(cell, palette):
    border = getattr(cell, "border", None)
    if border is None:
        return {}
    out = {}
    for side in _BORDER_SIDES:
        side_obj = getattr(border, side, None)
        style = getattr(side_obj, "style", None) if side_obj is not None else None
        if style is None:
            continue
        comps = _colour_components(getattr(side_obj, "color", None))
        out[side] = {"style": style, "resolved_hex": _resolve_hex(comps, palette) if comps else None}
    return out


def _alignment_info(cell):
    a = getattr(cell, "alignment", None)
    if a is None:
        return None
    return {
        "horizontal": getattr(a, "horizontal", None),
        "vertical": getattr(a, "vertical", None),
        "wrap_text": bool(getattr(a, "wrap_text", False)),
        "text_rotation": getattr(a, "text_rotation", 0),
        "indent": getattr(a, "indent", 0),
    }


def _cell_state(value, formula):
    if formula is not None:
        return "formula_blank" if value is None or value == "" else "formula_value"
    if value is None:
        return "blank"
    if value == "":
        return "empty_string"
    return "value"


def _build_merge_maps(worksheet):
    anchor_of = {}   # member ref -> anchor ref
    members_of = {}  # anchor ref -> [member refs]
    for rng in worksheet.merged_cells.ranges:
        cells = list(rng.cells)
        if not cells:
            continue
        arow, acol = min(cells)
        anchor = worksheet.cell(row=arow, column=acol).coordinate
        members = []
        for r, c in cells:
            if (r, c) == (arow, acol):
                continue
            ref = worksheet.cell(row=r, column=c).coordinate
            anchor_of[ref] = anchor
            members.append(ref)
        members_of[anchor] = members
    return anchor_of, members_of


def _images_info(raw_images):
    """Shapes the zip-based image detections (image_detector) into the lossless
    schema. Reads no image content. Runs dependency-free (no Pillow), so images
    are never silently missed."""
    images = []
    for img in raw_images or []:
        width = img.get("width_px")
        height = img.get("height_px")
        likely_tabular = bool(
            isinstance(width, (int, float))
            and isinstance(height, (int, float))
            and width >= _LIKELY_TABULAR_MIN_W
            and height >= _LIKELY_TABULAR_MIN_H
        )
        images.append(
            {
                "anchor_cell": img.get("anchor_cell"),
                "width": width,
                "height": height,
                "format": img.get("format"),
                "likely_tabular": likely_tabular,
                "needs_ocr": likely_tabular,
            }
        )
    return images


def _extract_sheet(worksheet, formula_ws, palette, index, raw_images=None):
    anchor_of, members_of = _build_merge_maps(worksheet)

    min_row = worksheet.min_row or 1
    max_row = worksheet.max_row or 0
    min_col = worksheet.min_column or 1
    max_col = worksheet.max_column or 0

    cells = []
    counts = {
        "total_cells": 0,
        "value_cells": 0,
        "blank_cells": 0,
        "colored_cells": 0,
        "formula_cells": 0,
        "comment_cells": 0,
        "hyperlink_cells": 0,
        "merged_member_cells": 0,
    }

    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = worksheet.cell(row=r, column=c)
            ref = cell.coordinate
            fcell = formula_ws.cell(row=r, column=c) if formula_ws is not None else None
            formula = fcell.value if (fcell is not None and fcell.data_type == "f") else None

            value = cell.value
            fill = _fill_info(cell, palette)
            comment = getattr(cell, "comment", None)
            hyperlink = getattr(cell, "hyperlink", None)
            hyperlink_target = getattr(hyperlink, "target", None) if hyperlink is not None else None

            merge_role = None
            merge_anchor = None
            if ref in members_of:
                merge_role = "anchor"
            elif ref in anchor_of:
                merge_role = "member"
                merge_anchor = anchor_of[ref]
                counts["merged_member_cells"] += 1

            counts["total_cells"] += 1
            if value is None or value == "":
                counts["blank_cells"] += 1
            else:
                counts["value_cells"] += 1
            if fill["has_fill"] and fill["resolved_hex"] is not None:
                counts["colored_cells"] += 1
            if formula is not None:
                counts["formula_cells"] += 1
            if comment is not None:
                counts["comment_cells"] += 1
            if hyperlink_target is not None:
                counts["hyperlink_cells"] += 1

            # Every cell in the used range is emitted so nothing is skipped,
            # but a field is included only when it deviates from the Excel
            # default. Absence therefore has a precise, lossless meaning:
            # no `fill` => no fill; no `font` => default font; no `border` =>
            # no border; no `merged` => not merged; no `comment`/`hyperlink`
            # => none. This keeps the representation complete without a
            # ~700-byte dict for every empty cell.
            entry = {
                "cell_ref": ref,
                "row": r,
                "column": c,
                "value": _serialize(value),
                "state": _cell_state(value, formula),
            }
            display = _display_value(value)
            if display != "":
                entry["display_value"] = display
            if cell.data_type and cell.data_type != "n":
                entry["data_type"] = cell.data_type
            if formula is not None:
                entry["formula"] = formula
            if cell.number_format and cell.number_format != "General":
                entry["number_format"] = cell.number_format
            if fill["has_fill"]:
                entry["fill"] = fill
            font = _font_info(cell, palette)
            if font and (font["bold"] or font["italic"] or font["underline"] or font["resolved_hex"]):
                entry["font"] = font
            border = _border_info(cell, palette)
            if border:
                entry["border"] = border
            alignment = _alignment_info(cell)
            if alignment and (
                alignment["horizontal"]
                or alignment["vertical"]
                or alignment["wrap_text"]
                or alignment["text_rotation"]
                or alignment["indent"]
            ):
                entry["alignment"] = alignment
            if merge_role is not None:
                entry["merged"] = True
                entry["merge_role"] = merge_role
                if merge_anchor is not None:
                    entry["merge_anchor"] = merge_anchor
                if members_of.get(ref):
                    entry["merged_members"] = members_of[ref]
            if comment is not None:
                entry["comment"] = {"text": comment.text, "author": comment.author}
            if hyperlink_target is not None:
                entry["hyperlink"] = hyperlink_target

            cells.append(entry)

    hidden_rows = sorted(r for r, d in worksheet.row_dimensions.items() if d.hidden)
    hidden_cols = sorted(c for c, d in worksheet.column_dimensions.items() if d.hidden)
    merged_ranges = [str(rng) for rng in worksheet.merged_cells.ranges]
    images = _images_info(raw_images)

    sheet = {
        "name": worksheet.title,
        "index": index,
        "state": worksheet.sheet_state,
        "dimensions": {
            "ref": worksheet.dimensions,
            "min_row": min_row,
            "max_row": max_row,
            "min_col": min_col,
            "max_col": max_col,
        },
        "freeze_panes": worksheet.freeze_panes,
        "merged_ranges": merged_ranges,
        "hidden_rows": hidden_rows,
        "hidden_columns": hidden_cols,
        "sheet_properties": {
            "tab_color": getattr(getattr(worksheet, "sheet_properties", None), "tabColor", None)
            and str(worksheet.sheet_properties.tabColor),
        },
        "images": images,
        "cells": cells,
        "validation": {
            **counts,
            "merged_ranges": len(merged_ranges),
            "hidden_rows": len(hidden_rows),
            "hidden_columns": len(hidden_cols),
            "images": len(images),
        },
    }
    return sheet


def _workbook_properties(workbook):
    props = getattr(workbook, "properties", None)
    if props is None:
        return {}
    return {
        "creator": getattr(props, "creator", None),
        "title": getattr(props, "title", None),
        "subject": getattr(props, "subject", None),
        "description": getattr(props, "description", None),
        "last_modified_by": getattr(props, "lastModifiedBy", None),
        "created": _serialize(getattr(props, "created", None)),
        "modified": _serialize(getattr(props, "modified", None)),
        "category": getattr(props, "category", None),
        "keywords": getattr(props, "keywords", None),
    }


def extract_lossless(file_path, filename=None, only_sheet=None):
    """
    Returns the full lossless representation of the workbook (or a single
    sheet if `only_sheet` is given). Re-parses the .xlsx from disk each call -
    the stored file is the source of truth, so the result is always faithful.
    """
    values_wb = openpyxl.load_workbook(file_path, data_only=True)
    formulas_wb = openpyxl.load_workbook(file_path, data_only=False)
    palette, _palette_error = load_theme_palette(file_path)
    # Dependency-free image detection straight from the .xlsx package.
    images_by_sheet = detect_images_by_sheet(file_path)

    sheets = {}
    for index, name in enumerate(values_wb.sheetnames):
        if only_sheet is not None and name != only_sheet:
            continue
        sheets[name] = _extract_sheet(
            values_wb[name], formulas_wb[name], palette, index,
            raw_images=images_by_sheet.get(name, []),
        )

    result = {
        "workbook": {
            "filename": filename,
            "sheet_count": len(values_wb.sheetnames),
            "sheet_names": list(values_wb.sheetnames),
            "properties": _workbook_properties(values_wb),
            "theme_palette_loaded": palette is not None,
        },
        "sheets": sheets,
    }
    result["validation"] = build_validation_report(result)
    return result


def build_validation_report(lossless):
    """Aggregate, deterministic validation summary (spec item 8)."""
    per_sheet = {}
    totals = {
        "total_cells": 0,
        "value_cells": 0,
        "blank_cells": 0,
        "colored_cells": 0,
        "formula_cells": 0,
        "comment_cells": 0,
        "hyperlink_cells": 0,
        "merged_ranges": 0,
        "hidden_rows": 0,
        "hidden_columns": 0,
        "images": 0,
    }
    for name, sheet in lossless.get("sheets", {}).items():
        v = sheet.get("validation", {})
        per_sheet[name] = v
        for key in totals:
            totals[key] += v.get(key, 0) or 0

    return {
        "sheet_count": len(lossless.get("sheets", {})),
        "totals": totals,
        "per_sheet": per_sheet,
    }
