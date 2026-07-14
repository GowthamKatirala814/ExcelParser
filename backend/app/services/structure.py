"""
Structural block/table detection.

Purely layout-driven: no knowledge of what any sheet, column, or color means.
Runs identically on any hotel workbook - nothing here references a specific
sheet name, room type, or column layout.
"""
import datetime

from openpyxl.utils import get_column_letter


def detect_blocks(worksheet):
    """
    Splits the worksheet's used range into row-range blocks. A new block
    starts whenever there are one or more fully blank rows (no value in any
    column of the used range) between two regions of non-blank rows.

    Returns a list of (start_row, end_row) tuples, inclusive, in row order.
    """
    min_row, min_col = worksheet.min_row, worksheet.min_column
    max_row, max_col = worksheet.max_row, worksheet.max_column
    if not max_row or not max_col:
        return []

    blocks = []
    current_start = None
    current_end = None

    for row in worksheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        row_num = row[0].row
        has_value = any(cell.value is not None for cell in row)
        if has_value:
            if current_start is None:
                current_start = row_num
            current_end = row_num
        else:
            if current_start is not None:
                blocks.append((current_start, current_end))
                current_start = None

    if current_start is not None:
        blocks.append((current_start, current_end))

    return blocks


def _row_pattern(worksheet, row_num, min_col, max_col):
    """Tuple of bool(has value) per column - the row's visual "shape"."""
    return tuple(
        worksheet.cell(row=row_num, column=c).value is not None
        for c in range(min_col, max_col + 1)
    )


def _row_is_bold(worksheet, row_num, min_col, max_col):
    for c in range(min_col, max_col + 1):
        cell = worksheet.cell(row=row_num, column=c)
        if cell.value is None:
            continue
        font = getattr(cell, "font", None)
        if font is not None and getattr(font, "bold", False):
            return True
    return False


def _row_has_horizontal_merge(worksheet, row_num):
    for merged_range in worksheet.merged_cells.ranges:
        if (
            merged_range.min_row == row_num
            and merged_range.max_row == row_num
            and (merged_range.max_col - merged_range.min_col + 1) >= 2
        ):
            return True
    return False


def detect_header(worksheet, block_start, block_end, min_col, max_col):
    """
    Determines whether the first row of a block is a header row, using three
    concrete signals (bold cells, a horizontal merge across >=2 columns, or
    >=2 immediately-following rows that repeat the same non-empty-column
    shape). Returns (header_row_or_None, signals_dict_or_None).
    """
    first_row = block_start

    is_bold = _row_is_bold(worksheet, first_row, min_col, max_col)
    is_merged = _row_has_horizontal_merge(worksheet, first_row)

    repeats = False
    r1, r2 = first_row + 1, first_row + 2
    if r2 <= block_end:
        p1 = _row_pattern(worksheet, r1, min_col, max_col)
        p2 = _row_pattern(worksheet, r2, min_col, max_col)
        if p1 == p2 and any(p1):
            repeats = True

    if is_bold or is_merged or repeats:
        return first_row, {
            "bold": is_bold,
            "merged_across_columns": is_merged,
            "repeated_data_row_pattern": repeats,
        }

    return None, None


def _cell_for_structured(worksheet, row_num, col_num, raw_by_ref, member_to_anchor):
    """
    Builds one cell of the structured view, always traceable back to a raw
    cell-level entry: either its own (raw_by_ref[cell_ref]) or, for a
    non-anchor merged cell, the anchor's - whose value/color it inherits so
    the merged value is associated with every column it visually spans.
    """
    ref = worksheet.cell(row=row_num, column=col_num).coordinate

    if ref in member_to_anchor:
        anchor_ref = member_to_anchor[ref]
        anchor_raw = raw_by_ref.get(anchor_ref)
        return {
            "cell_ref": ref,
            "raw_cell_ref": anchor_ref,
            "value": anchor_raw.get("value") if anchor_raw else None,
            "color": anchor_raw.get("color") if anchor_raw else None,
            "merged_from": anchor_ref,
            "human_label": anchor_raw.get("human_label") if anchor_raw else None,
        }

    raw = raw_by_ref.get(ref)
    return {
        "cell_ref": ref,
        "raw_cell_ref": ref,
        "value": raw.get("value") if raw else None,
        "color": raw.get("color") if raw else None,
        "merged_from": None,
        "human_label": raw.get("human_label") if raw else None,
    }


def _find_date_row(worksheet, start, end, min_col, max_col):
    """
    A row within the block that establishes a horizontal date header, with at
    least one row below it in the block (i.e. it sits directly above data
    rows). Two concrete, purely positional patterns are recognized:

    1. Every populated cell in the row is a date/datetime, with >=2 such
       cells - a horizontal date header ("column E = 2026-06-01, column F =
       2026-06-02..."). Requiring >=2 columns distinguishes this from a
       single date value that happens to be the only entry in an otherwise
       blank row (e.g. one row per travel date in a vertical list), which is
       not a column index at all.
    2. Exactly one populated cell is a date (a month/period anchor) and the
       rest of the populated cells are the exact integer sequence 1, 2, 3...
       (a day-of-month index) - a common calendar-grid layout where the
       month is given once and each column is simply "day N" of it. Reading
       day N of the anchor's month is a positional fact, not a guess.

    Returns (date_row_or_None, anchor_date_or_None). anchor_date is only set
    for pattern 2.
    """
    for r in range(start, end):
        values = [worksheet.cell(row=r, column=c).value for c in range(min_col, max_col + 1)]
        present = [v for v in values if v is not None]
        if len(present) >= 2 and all(isinstance(v, (datetime.datetime, datetime.date)) for v in present):
            return r, None

    for r in range(start, end):
        populated = [
            worksheet.cell(row=r, column=c).value
            for c in range(min_col, max_col + 1)
            if worksheet.cell(row=r, column=c).value is not None
        ]
        if len(populated) < 3:
            continue
        date_values = [v for v in populated if isinstance(v, (datetime.datetime, datetime.date))]
        int_values = [v for v in populated if isinstance(v, int) and not isinstance(v, bool)]
        if len(date_values) == 1 and len(int_values) == len(populated) - 1:
            if int_values == list(range(1, len(int_values) + 1)):
                return r, date_values[0]

    return None, None


def _as_date(value):
    """Normalizes a datetime.datetime or datetime.date to a plain date - the
    column_index only ever represents day-level granularity, and a stray
    time-of-day component (e.g. midnight from an Excel date-only cell) must
    never leak into a "date" string, since downstream code parses these with
    date.fromisoformat(), which rejects a datetime-with-time string."""
    if isinstance(value, datetime.datetime):
        return value.date()
    return value


def _build_column_index(worksheet, date_row, anchor_date, min_col, max_col):
    """Maps each date-bearing column letter to its date and day number."""
    index = {}

    if anchor_date is not None:
        anchor_date = _as_date(anchor_date)
        # Anchor + day-index pattern: the anchor cell itself is the
        # month/period label, not a day column - only the integer-valued
        # cells are actual day slots, and the sheet's own integer is the
        # ground-truth day number (not a recount).
        for c in range(min_col, max_col + 1):
            value = worksheet.cell(row=date_row, column=c).value
            if not isinstance(value, int) or isinstance(value, bool):
                continue
            try:
                date_value = anchor_date.replace(day=value)
            except ValueError:
                # day-of-month out of range for the anchor's month - be
                # honest rather than guess a rollover into the next month.
                date_value = None
            index[get_column_letter(c)] = {
                "date": date_value.isoformat() if date_value is not None else None,
                "day_number": value,
            }
        return index

    day_number = 0
    for c in range(min_col, max_col + 1):
        value = _as_date(worksheet.cell(row=date_row, column=c).value)
        if value is None:
            continue
        day_number += 1
        index[get_column_letter(c)] = {
            "date": value.isoformat() if hasattr(value, "isoformat") else str(value),
            "day_number": day_number,
        }
    return index


def _combine_reason(existing, new_reason):
    if not existing:
        return new_reason
    if new_reason in existing:
        return existing
    return f"{existing}; {new_reason}"


def _mark_adjacent_block_ambiguity(worksheet, tables, min_col, max_col):
    """
    Flags two adjacent blocks as possibly-one-table when they're separated by
    exactly a single blank row AND the row bordering the gap on each side
    shares the same populated-column shape - a concrete, checkable signal
    that the "gap" may be a spacer row inside one table rather than a true
    boundary between two tables. Never merges them; only flags the doubt.
    """
    for i in range(len(tables) - 1):
        current, following = tables[i], tables[i + 1]
        gap = following["row_range"][0] - current["row_range"][1] - 1
        if gap != 1:
            continue

        last_row_of_current = current["row_range"][1]
        first_row_of_following = following["row_range"][0]
        pattern_before = _row_pattern(worksheet, last_row_of_current, min_col, max_col)
        pattern_after = _row_pattern(worksheet, first_row_of_following, min_col, max_col)

        if pattern_before == pattern_after and any(pattern_before):
            reason = (
                f"blocks at rows {current['row_range']} and {following['row_range']} are "
                "separated by exactly one blank row and share the same populated-column "
                "layout; they may be a single table split by a spacer row rather than two "
                "distinct tables"
            )
            current["ambiguous"] = True
            current["ambiguous_reason"] = _combine_reason(current["ambiguous_reason"], reason)
            following["ambiguous"] = True
            following["ambiguous_reason"] = _combine_reason(following["ambiguous_reason"], reason)


def build_tables(worksheet, raw_by_ref, member_to_anchor):
    """
    Assembles the sheet's blocks into tables: row-range, detected header (or
    None), and rows of fully-linked cells. This is pure assembly on top of
    detect_blocks()/detect_header() - no new detection logic here.
    """
    min_col, max_col = worksheet.min_column, worksheet.max_column
    tables = []

    for start, end in detect_blocks(worksheet):
        header_row, signals = detect_header(worksheet, start, end, min_col, max_col)
        data_row_start = start + 1 if header_row is not None else start

        def build_row(row_num):
            return {
                "row": row_num,
                "cells": [
                    _cell_for_structured(worksheet, row_num, c, raw_by_ref, member_to_anchor)
                    for c in range(min_col, max_col + 1)
                ],
            }

        header = None
        if header_row is not None:
            header = build_row(header_row)
            header["signals"] = signals

        data_rows = [build_row(r) for r in range(data_row_start, end + 1)]

        date_row, anchor_date = _find_date_row(worksheet, start, end, min_col, max_col)
        column_index = (
            _build_column_index(worksheet, date_row, anchor_date, min_col, max_col)
            if date_row is not None
            else None
        )

        ambiguous = False
        ambiguous_reason = None
        if header_row is None and end > start:
            # A block with more than one row but no bold/merge/repeated-shape
            # signal genuinely could not be classified as headered or
            # headerless with confidence - flagged rather than guessed.
            ambiguous = True
            ambiguous_reason = (
                f"block at rows {start}-{end} has no bold, merged-header, or "
                "repeated-data-row-pattern signal on its first row, so it's unclear whether "
                "it has a header at all; all rows are being treated as data"
            )

        tables.append(
            {
                "row_range": [start, end],
                "header": header,
                "rows": data_rows,
                "date_row": date_row,
                "column_index": column_index,
                "ambiguous": ambiguous,
                "ambiguous_reason": ambiguous_reason,
            }
        )

    _mark_adjacent_block_ambiguity(worksheet, tables, min_col, max_col)

    return tables
