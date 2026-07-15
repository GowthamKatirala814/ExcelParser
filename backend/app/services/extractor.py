"""
Per-workbook, per-sheet extraction. Sheets are always processed independently
and never merged together. Every statistic reported here is a direct count of
something that was actually read from the file - nothing is estimated.
"""
import datetime
import time

import openpyxl

from .cf_resolver import resolve_conditional_format_color
from .image_detector import detect_images_by_sheet
from .color_resolver import (
    resolve_color_object,
    resolve_column_default_fill,
    resolve_direct_fill,
    resolve_row_default_fill,
)
from .structure import build_tables
from .theme_resolver import load_theme_palette

MAX_INVENTORY_EXAMPLES = 5
_BORDER_SIDES = ("top", "bottom", "left", "right")

# Perceptual (CIE76 Delta-E in LAB space) distance below which two fill colors
# are generally very hard for a human reviewer to tell apart. 5 is a common,
# defensible "just noticeable difference" boundary; below it a person is
# likely to read the two colors as the same. Purely a flag for human review -
# nothing in the pipeline ever merges or renames these colors.
LOW_CONTRAST_DELTA_E_THRESHOLD = 5.0
_WHITE_HEX = "FFFFFF"


def _serialize_value(value):
    if isinstance(value, (datetime.datetime, datetime.date, datetime.time)):
        return value.isoformat()
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return str(value)


def _resolve_cell_color(cell, cell_value, worksheet, workbook, theme_palette, theme_error):
    """
    Resolution precedence matches Excel's own: explicit per-cell fill first,
    then (only when the cell has no style of its own) the row default, then the
    column default, then conditional formatting.
    """
    direct = resolve_direct_fill(cell, theme_palette, theme_error)
    if direct is not None:
        direct["source"] = "direct_fill"
        return direct

    # Only fall back to row/column defaults when the cell has no explicit style
    # of its own - an explicit per-cell style always wins in Excel, even one
    # that resolves to "no fill". Row default beats column default.
    if getattr(cell, "style_id", None) == 0:
        row_default = resolve_row_default_fill(cell, worksheet, workbook, theme_palette, theme_error)
        if row_default is not None:
            row_default["source"] = "row_default"
            return row_default

        col_default = resolve_column_default_fill(cell, worksheet, workbook, theme_palette, theme_error)
        if col_default is not None:
            col_default["source"] = "column_default"
            return col_default

    cf_result = resolve_conditional_format_color(
        cell, cell_value, worksheet, workbook, theme_palette, theme_error
    )
    if cf_result is not None:
        cf_result["source"] = "conditional_formatting"
        return cf_result

    return None  # no fill at all: blank / no-fill cell


def _resolve_font_color(cell, theme_palette, theme_error):
    font = getattr(cell, "font", None)
    color = getattr(font, "color", None) if font is not None else None
    if color is None:
        return None
    return resolve_color_object(color, theme_palette, theme_error)


def _resolve_borders(cell, theme_palette, theme_error):
    """Returns {side: {"style": str, "color": color-fact}} for sides that actually have a border."""
    border = getattr(cell, "border", None)
    if border is None:
        return {}
    result = {}
    for side in _BORDER_SIDES:
        side_border = getattr(border, side, None)
        style = getattr(side_border, "style", None) if side_border is not None else None
        if style is None:
            continue
        result[side] = {
            "style": style,
            "color": resolve_color_object(
                getattr(side_border, "color", None), theme_palette, theme_error
            ),
        }
    return result


def _resolve_comment(cell):
    comment = getattr(cell, "comment", None)
    if comment is None:
        return None
    return {"text": comment.text, "author": comment.author}


def _resolve_formula(formula_cell):
    """formula_cell comes from a parallel data_only=False load; None if not applicable."""
    if formula_cell is None or formula_cell.data_type != "f":
        return None
    return formula_cell.value


def _inventory_key(color_result):
    if color_result is None:
        return None
    if color_result.get("resolved"):
        return f"#{color_result['hex']}"
    if color_result.get("theme_ref"):
        return color_result["theme_ref"]
    return f"unresolved:{color_result.get('reason_unresolved', 'unknown')}"


# --- Low-contrast / faded color detection (perceptual, LAB Delta-E) --------
#
# Two fills can differ by only a few RGB units yet look identical to a person,
# or a fill can be so close to white it's effectively invisible. RGB
# subtraction does not match human perception, so distance is measured as
# CIE76 Delta-E in CIELAB space (sRGB -> linear -> XYZ(D65) -> LAB). This is a
# descriptive flag only: nothing downstream merges, renames, or equates the
# flagged colors.


def _hex_to_lab(hex6):
    r = int(hex6[0:2], 16) / 255.0
    g = int(hex6[2:4], 16) / 255.0
    b = int(hex6[4:6], 16) / 255.0

    def _linear(c):
        return ((c + 0.055) / 1.055) ** 2.4 if c > 0.04045 else c / 12.92

    r, g, b = _linear(r), _linear(g), _linear(b)
    # linear sRGB -> XYZ (D65)
    x = r * 0.4124 + g * 0.3576 + b * 0.1805
    y = r * 0.2126 + g * 0.7152 + b * 0.0722
    z = r * 0.0193 + g * 0.1192 + b * 0.9505
    # normalize by D65 reference white
    x, y, z = x / 0.95047, y / 1.0, z / 1.08883

    def _f(t):
        return t ** (1.0 / 3.0) if t > 0.008856 else 7.787 * t + 16.0 / 116.0

    fx, fy, fz = _f(x), _f(y), _f(z)
    return (116.0 * fy - 16.0, 500.0 * (fx - fy), 200.0 * (fy - fz))


def _delta_e_cie76(lab1, lab2):
    return ((lab1[0] - lab2[0]) ** 2 + (lab1[1] - lab2[1]) ** 2 + (lab1[2] - lab2[2]) ** 2) ** 0.5


def _detect_low_contrast_pairs(color_inventory):
    """
    Compares every pair of distinct RESOLVED colors in the inventory (and each
    against pure white) and flags any whose perceptual distance is below the
    threshold. Returns a list of descriptive dicts.
    """
    resolved = [
        {"hex": e["hex_or_theme_ref"], "count": e.get("cell_count", 0), "lab": _hex_to_lab(e["hex_or_theme_ref"].lstrip("#"))}
        for e in color_inventory
        if e.get("resolved") and isinstance(e.get("hex_or_theme_ref"), str) and e["hex_or_theme_ref"].startswith("#")
    ]
    white_lab = _hex_to_lab(_WHITE_HEX)
    pairs = []

    for i in range(len(resolved)):
        a = resolved[i]
        # near-white (effectively invisible) fills, excluding white itself
        if a["hex"].lstrip("#").upper() != _WHITE_HEX:
            de_white = _delta_e_cie76(a["lab"], white_lab)
            if de_white < LOW_CONTRAST_DELTA_E_THRESHOLD:
                pairs.append(
                    {
                        "color_a": a["hex"],
                        "color_b": f"#{_WHITE_HEX}",
                        "delta_e": round(de_white, 2),
                        "cell_count_a": a["count"],
                        "cell_count_b": None,
                        "note": "this color is nearly white and may be visually indistinguishable from no fill",
                    }
                )
        for j in range(i + 1, len(resolved)):
            b = resolved[j]
            de = _delta_e_cie76(a["lab"], b["lab"])
            if de < LOW_CONTRAST_DELTA_E_THRESHOLD:
                pairs.append(
                    {
                        "color_a": a["hex"],
                        "color_b": b["hex"],
                        "delta_e": round(de, 2),
                        "cell_count_a": a["count"],
                        "cell_count_b": b["count"],
                        "note": "these two colors may be visually indistinguishable to a human reviewer",
                    }
                )
    return pairs


# Embedded image detection lives in image_detector.detect_images_by_sheet -
# it reads the .xlsx drawing parts directly (no openpyxl/Pillow), so images
# are never silently missed. extract_workbook computes it once and passes each
# sheet's list into extract_sheet. Location/size only, never image content.


def _build_extraction_summary(rows, unresolved_cells, embedded_images, low_contrast_pairs, worksheet):
    """
    A self-auditing rollup so a human reviewing the extraction knows exactly
    where to double-check rather than trusting the raw cell dump blind: how
    each resolved cell's color was actually determined (own explicit fill vs.
    row default vs. column default vs. conditional formatting), plus counts of
    anything advisory (unresolved colors, near-identical colors, images).
    Purely descriptive - it reads the already-produced rows, changing nothing.
    """
    explicit_fill = row_default = column_default = conditional_formatting = 0
    for row in rows:
        color = row.get("color") or {}
        if not color.get("resolved"):
            continue
        source = color.get("source")
        if source == "direct_fill":
            explicit_fill += 1
        elif source == "row_default":
            row_default += 1
        elif source == "column_default":
            column_default += 1
        elif source == "conditional_formatting":
            conditional_formatting += 1

    warnings = []
    if embedded_images:
        cells = ", ".join(img.get("anchor_cell") or "unknown position" for img in embedded_images)
        warnings.append(
            f"{len(embedded_images)} embedded image(s) detected at {cells} - "
            "their content was not read; check manually if they contain data"
        )
    if unresolved_cells:
        warnings.append(
            f"{unresolved_cells} cell(s) have a color that could not be resolved to an exact value"
        )
    if low_contrast_pairs:
        warnings.append(
            f"{len(low_contrast_pairs)} pair(s) of near-identical colors detected - "
            "may be visually indistinguishable to a human reviewer"
        )

    return {
        # Number of reported cell rows (excludes non-anchor merged members);
        # the sheet's full scanned count is the top-level "total_cells".
        "cells_reported": len(rows),
        "cells_with_explicit_fill": explicit_fill,
        "cells_using_row_default": row_default,
        "cells_using_column_default": column_default,
        "cells_using_conditional_formatting": conditional_formatting,
        "unresolved_colors": unresolved_cells,
        "low_contrast_color_pairs": len(low_contrast_pairs),
        "embedded_images": len(embedded_images),
        "merged_ranges": len(worksheet.merged_cells.ranges),
        "warnings": warnings,
    }


def _build_merge_maps(worksheet):
    anchor_to_members = {}
    member_to_anchor = {}
    for merged_range in worksheet.merged_cells.ranges:
        cells = list(merged_range.cells)
        if not cells:
            continue
        anchor_row, anchor_col = min(cells)
        anchor_cell = worksheet.cell(row=anchor_row, column=anchor_col)
        anchor_ref = anchor_cell.coordinate
        members = []
        for row, col in cells:
            if (row, col) == (anchor_row, anchor_col):
                continue
            ref = worksheet.cell(row=row, column=col).coordinate
            members.append(ref)
            member_to_anchor[ref] = anchor_ref
        anchor_to_members[anchor_ref] = members
    return anchor_to_members, member_to_anchor


def extract_sheet(
    workbook,
    worksheet,
    theme_palette=None,
    theme_error=None,
    formula_worksheet=None,
    embedded_images=None,
):
    anchor_to_members, member_to_anchor = _build_merge_maps(worksheet)

    total_cells = 0
    colored_cells = 0
    blank_cells = 0
    unresolved_cells = 0
    ambiguous_cells = 0

    inventory = {}  # key -> {hex_or_theme_ref, resolved, cell_count, example_refs, reason_unresolved}
    rows = []

    for row in worksheet.iter_rows():
        for cell in row:
            ref = cell.coordinate
            if cell.value is None and ref in member_to_anchor:
                # non-anchor merged cell: still scanned for coverage stats,
                # but not reported as its own row (its value belongs to the anchor).
                total_cells += 1
                raw_value = None
            else:
                total_cells += 1
                raw_value = cell.value

            color_result = _resolve_cell_color(
                cell, raw_value, worksheet, workbook, theme_palette, theme_error
            )

            if color_result is None:
                blank_cells += 1
            elif color_result.get("resolved"):
                colored_cells += 1
            else:
                unresolved_cells += 1
                if color_result.get("ambiguous"):
                    ambiguous_cells += 1

            if color_result is not None:
                key = _inventory_key(color_result)
                entry = inventory.setdefault(
                    key,
                    {
                        "key": key,
                        "hex_or_theme_ref": color_result.get("hex")
                        and f"#{color_result['hex']}"
                        or color_result.get("theme_ref")
                        or key,
                        "resolved": bool(color_result.get("resolved")),
                        "reason_unresolved": color_result.get("reason_unresolved"),
                        "cell_count": 0,
                        "example_refs": [],
                    },
                )
                entry["cell_count"] += 1
                if len(entry["example_refs"]) < MAX_INVENTORY_EXAMPLES:
                    entry["example_refs"].append(ref)

            if ref in member_to_anchor:
                # Skip emitting a separate row for non-anchor merged members.
                continue

            formula_cell = formula_worksheet[ref] if formula_worksheet is not None else None

            rows.append(
                {
                    "cell_ref": ref,
                    "value": _serialize_value(cell.value),
                    "color": {
                        "resolved": bool(color_result and color_result.get("resolved")),
                        "hex": color_result.get("hex") if color_result else None,
                        "theme_ref": color_result.get("theme_ref") if color_result else None,
                        "source": color_result.get("source") if color_result else None,
                        "ambiguous": bool(color_result and color_result.get("ambiguous")),
                        "reason_unresolved": color_result.get("reason_unresolved")
                        if color_result
                        else None,
                    },
                    "font_color": _resolve_font_color(cell, theme_palette, theme_error),
                    "borders": _resolve_borders(cell, theme_palette, theme_error),
                    "comment": _resolve_comment(cell),
                    "formula": _resolve_formula(formula_cell),
                    "merged_with": anchor_to_members.get(ref, []),
                    "human_label": None,
                }
            )

    color_inventory = sorted(inventory.values(), key=lambda e: e["cell_count"], reverse=True)

    # Purely descriptive extra passes - they never alter any value, cell, or
    # color resolution above; they only add advisory flags for human review.
    low_contrast_pairs = _detect_low_contrast_pairs(color_inventory)
    embedded_images = embedded_images or []
    extraction_summary = _build_extraction_summary(
        rows, unresolved_cells, embedded_images, low_contrast_pairs, worksheet
    )

    raw_cells = {
        "total_cells": total_cells,
        "colored_cells": colored_cells,
        "blank_cells": blank_cells,
        "unresolved_cells": unresolved_cells,
        "ambiguous_cells": ambiguous_cells,
        "color_inventory": color_inventory,
        "rows": rows,
    }

    raw_by_ref = {row["cell_ref"]: row for row in rows}
    structured = {"tables": build_tables(worksheet, raw_by_ref, member_to_anchor)}

    return {
        # Flat fields kept for backward compatibility with existing callers
        # (routers, reports) - never removed or degraded.
        "total_cells": total_cells,
        "colored_cells": colored_cells,
        "blank_cells": blank_cells,
        "unresolved_cells": unresolved_cells,
        "ambiguous_cells": ambiguous_cells,
        "color_inventory": color_inventory,
        "rows": rows,
        # New: explicit raw/structured split, side by side, per the
        # workbook -> sheets -> tables -> rows -> cells hierarchy.
        "raw_cells": raw_cells,
        "structured": structured,
        # New advisory-only flags (isolated; nothing above depends on them).
        "low_contrast_pairs": low_contrast_pairs,
        "embedded_images": embedded_images,
        "extraction_summary": extraction_summary,
    }


def extract_workbook(file_path):
    """
    Returns (per_sheet_results: dict[sheet_name -> extraction result with
    processing_time_seconds], workbook_total_time_seconds).
    """
    workbook_start = time.perf_counter()
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    # A second, parallel load with formulas intact (data_only=True replaces
    # formula cells with their cached computed value, so formula text can
    # only be read from a workbook opened without that flag).
    formula_workbook = openpyxl.load_workbook(file_path, data_only=False)
    theme_palette, theme_error = load_theme_palette(file_path)
    # Detected once from the .xlsx package directly (dependency-free), then
    # handed to each sheet - never via openpyxl/Pillow, so images are not
    # silently missed when Pillow is absent.
    images_by_sheet = detect_images_by_sheet(file_path)

    per_sheet_results = {}
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        formula_worksheet = formula_workbook[sheet_name]
        sheet_start = time.perf_counter()
        result = extract_sheet(
            workbook,
            worksheet,
            theme_palette,
            theme_error,
            formula_worksheet,
            embedded_images=images_by_sheet.get(sheet_name, []),
        )
        result["processing_time_seconds"] = time.perf_counter() - sheet_start
        per_sheet_results[sheet_name] = result

    workbook_total_time = time.perf_counter() - workbook_start
    return per_sheet_results, workbook_total_time


def read_sheet_dimensions(file_path):
    """Used right after upload, before extraction, for immediate metadata."""
    workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    dimensions = {}
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        dimensions[sheet_name] = {
            "rows": worksheet.max_row or 0,
            "columns": worksheet.max_column or 0,
        }
    workbook.close()
    return workbook.sheetnames, dimensions
